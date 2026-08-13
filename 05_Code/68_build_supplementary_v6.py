# ============================================================
# 68_build_supplementary_v6.py —— 补充表工作簿 S1–S12（v7 口径，含 v6 合并参考页）
# 依据：修改报告_BaP_TBI_最终版.md D-1（S1=直系同源、S2=置换、S3=敏感性…）
# 输出：submission/03_Supplementary/Tables/Supplementary_Tables_S1-S12.xlsx
# ============================================================
import csv, os
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ROOT = os.getcwd()
OUT = os.path.join(ROOT, "submission", "03_Supplementary", "Tables",
                   "Supplementary_Tables_S1-S12.xlsx")
O = os.path.join(ROOT, "results", "16_Orthology")
R = os.path.join(ROOT, "results")

def csv_to_rows(path):
    with open(path, encoding="utf-8") as f:
        return list(csv.reader(f))

def md_to_rows(path):
    with open(path, encoding="utf-8") as f:
        return [[line] for line in f.read().splitlines()]

wb = Workbook()
wb.remove(wb.active)

sheets = [
    ("S1_Orthology_Mapping",
     os.path.join(O, "orthology_mapping_full.csv"), "Ensembl Genes 116; 364 human targets; 1:1/1:many; Hprt1->Hprt alias resolved"),
    ("S2_Permutation",
     os.path.join(O, "v7_single_cohort", "permutation_v7.csv"), "v7 primary (GSE58485-only): seed=2026; 10,000 draws; union (K=11 vs 364) and best-module (K=3 vs 109); v6 merged reference in permutation_results.csv"),
    ("S3_Sensitivity_Matrix",
     os.path.join(O, "sensitivity_matrix.csv"), "CLEAN/FULL/GSE58485-only/RBE/SVA/no-ComBat; orthology-based candidates"),
    ("S4_ML_Benchmark",
     os.path.join(R, "07_ML_Signature", "D3_ml_benchmark_v7", "model_comparison.csv"), "v7: 11 algorithms, 11 genes, n=24, 5x repeated 5-fold CV"),
    ("S4_ENet_Calibration",
     os.path.join(R, "07_ML_Signature", "D3_ml_benchmark_v7", "calibration_stats.csv"), "v7 ENet alpha=0.9 (perfect separation; descriptive only)"),
    ("S4_ENet_OOF",
     os.path.join(R, "07_ML_Signature", "D3_ml_benchmark_v7", "oof_auc_ci.csv"), "v7 ENet OOF AUC 95% CI"),
    ("S5_GO_AhR",
     os.path.join(R, "06_GO_KEGG_Enrichment", "v7_single_cohort", "GO_ALL_keep.csv"), "v7 GO BP terms (38, Count>=3); KEGG 1; AhR coverage not evaluable (no Fisher P)"),
    ("S6_Welch_Assumptions",
     os.path.join(R, "08_External_Validation", "assumption_checks.csv"), "Shapiro-Wilk + F-test per gene-cohort"),
    ("S7_snRNA_Composition",
     os.path.join(R, "11_SingleCell", "celltype_composition_stats.csv"), "sample-level composition stats"),
    ("S8_Docking",
     os.path.join(R, "10_Molecular_Docking", "04_docking", "docking_energies.csv"), "AutoDock Vina scores"),
    ("S9_Astro",
     os.path.join(R, "14_Astrocyte_Subclusters", "D8_astro", "astro_subcluster_group_stats.csv"), "astro subcluster stats"),
    ("S10_PositiveControl",
     os.path.join(O, "geo_positive_control_search.csv"), "GSE75206 analyzed; 52-series search"),
    ("S11_Deconvolution",
     os.path.join(R, "09_Immune_Deconvolution", "deconvolution_method_corr.csv"), "method correlation"),
]

header_font = Font(bold=True)
for name, path, note in sheets:
    if not os.path.exists(path):
        ws = wb.create_sheet(name[:31])
        ws.append([f"MISSING FILE: {path}"])
        continue
    ws = wb.create_sheet(name[:31])
    ws.append([note])
    ws.append([])
    for row in csv_to_rows(path):
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(italic=True, color="666666")
    for cell in ws[3]:
        cell.font = header_font
    for col_idx in range(1, min(len(ws[1]), 40) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 24

# v7 主分析补充页（敏感性矩阵附页 + 候选统计 + 方向复制 + GSE75206 结果）
v7dir = os.path.join(O, "v7_single_cohort")
extra_sheets = [
    ("S3b_V7_Primary",
     os.path.join(v7dir, "candidate_stats_GSE58485_only.csv"),
     "v7 primary (GSE58485-only): 11 candidates, logFC/P/adjP"),
    ("S6b_V7_Direction",
     os.path.join(R, "08_External_Validation", "v7_single_cohort",
                  "validation_direction_v7.csv"),
     "v7 direction replication (GSE128543/GSE205958)"),
    ("S10b_GSE75206",
     os.path.join(O, "positive_control_GSE75206", "candidate_deg_intersection.csv"),
     "GSE75206: evaluable-target DEG intersection (0)"),
    ("S12_CompoundControl",
     os.path.join(O, "v7_compound_control", "compound_control_results.csv"),
     "Non-BaP compound control (ChEMBL single-source; pyrene/phenanthrene); "
     "0 evaluable targets after 1:1 orthology + coverage filter"),
]
for name, path, note in extra_sheets:
    if not os.path.exists(path):
        ws = wb.create_sheet(name[:31])
        ws.append([f"MISSING FILE: {path}"])
        continue
    ws = wb.create_sheet(name[:31])
    ws.append([note])
    ws.append([])
    for row in csv_to_rows(path):
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(italic=True, color="666666")
    for cell in ws[3]:
        cell.font = header_font

# 评审回应新增页（BMC 投稿前评审 M1–M5 / m8–m9）
review = os.path.join(O, "review_supplements")
review_sheets = [
    ("S2b_MeanVar_Permutation",
     os.path.join(review, "S2_permutation_v7_meanvar.csv"),
     "Mean-variance-decile-matched permutation (primary 5,043 matrix; seed=2026; 10,000 draws)"),
    ("S2c_Strata_Counts",
     os.path.join(review, "permutation_strata_counts.csv"),
     "Gene counts per mean x variance decile cell (primary matrix)"),
    ("S3_WGCNA_Diagnostics",
     os.path.join(v7dir, "wgcna_soft_threshold_GSE58485_only.csv"),
     "Soft-threshold fit table (primary matrix; softPower 4, SFT R2=0.90)"),
    ("S3_WGCNA_Module_Sizes",
     os.path.join(v7dir, "wgcna_module_sizes_GSE58485_only.csv"),
     "Module sizes (primary matrix; green 256; union of significant modules 1,942)"),
    ("S6c_V7_CI",
     os.path.join(review, "S6_validation_ci_v7.csv"),
     "Per-gene Welch 95% CI for direction consistency (11 candidates x 2 cohorts)"),
    ("S10c_GSE75206_Power",
     os.path.join(review, "power_GSE75206.csv"),
     "Minimal detectable log2 fold change, n=4 vs 4, alpha=0.05, power=0.80"),
    ("S13_Sample_Manifest",
     os.path.join(review, "S13_sample_manifest.csv"),
     "All 47 samples: cohort/group/genotype/treatment/inclusion flags/exclusion reasons"),
    ("S14_FullUniverse_Summary",
     os.path.join(review, "S14_full_universe_summary.csv"),
     "Full-platform (GPL1261, 27,250 genes) sensitivity analysis summary"),
    ("S14_FullUniverse_Candidates",
     os.path.join(review, "S14_full_universe_candidates.csv"),
     "Full-platform candidates (6, all tier C)"),
    ("S14_FullUniverse_AhR",
     os.path.join(review, "S14_full_universe_ahr_battery.csv"),
     "AhR battery status in the full-platform universe"),
    ("S14_FullUniverse_Permutation",
     os.path.join(review, "S14_full_universe_permutation.csv"),
     "Full-platform permutation (uniform + mean-variance stratified)"),
    ("S14_FullUniverse_Validation",
     os.path.join(review, "S14_full_universe_validation.csv"),
     "Direction consistency of full-platform candidates (11/12)"),
]
for name, path, note in review_sheets:
    if not os.path.exists(path):
        ws = wb.create_sheet(name[:31])
        ws.append([f"MISSING FILE: {path}"])
        continue
    ws = wb.create_sheet(name[:31])
    ws.append([note])
    ws.append([])
    for row in csv_to_rows(path):
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(italic=True, color="666666")
    for cell in ws[3]:
        cell.font = header_font

# 追加 S6 hub 基因 DEG 统计（与手稿补充表 S6 内容一致；Akt1/Tnf 未覆盖则标 NA）
hub_path = os.path.join(R, "08_External_Validation", "hub_genes_deg.csv")
if os.path.exists(hub_path):
    ws6 = wb["S6_Welch_Assumptions"]
    ws6.append([])
    ws6.append(["--- hub-gene DEG statistics from prior study [2] ---"])
    for row in csv_to_rows(hub_path):
        ws6.append(row)

# 阈值扫描汇总（S3 附页）
ws = wb.create_sheet("S3_Threshold_Scan")
ws.append(["threshold", "path", "genes_after_filter", "DEGs", "candidates", "candidate_genes"])
for t in [0, 0.05, 0.1, 0.15, 0.2, 0.3, 0.5]:
    tag = ("%.2f" % t).replace(".", "")
    for suffix, label in [("", "CLEAN"), ("_FULL", "FULL")]:
        candf = os.path.join(R, "05_Candidate_Intersection", f"candidate_genes{suffix}_COV{tag}.csv")
        degf = os.path.join(R, "03_DEG_Analysis", f"deg_significant{suffix}_COV{tag}.csv")
        exf = os.path.join(R, "00_Data_Prep", f"corrected_expr{suffix}_COV{tag}.csv")
        cand = []
        if os.path.exists(candf):
            cand = [r[0] for r in csv_to_rows(candf)[1:]]
        ndeg = -1
        if os.path.exists(degf):
            ndeg = len(csv_to_rows(degf)) - 1
        ngenes = -1
        if os.path.exists(exf):
            ngenes = len(csv_to_rows(exf)) - 1
        ws.append([t, label, ngenes, ndeg, len(cand), ",".join(cand)])

wb.save(OUT)
print("written:", OUT)
